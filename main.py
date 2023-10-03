from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import openpyxl

def close_btn():
    try:
        close_btn = driver.find_element(By.XPATH, "//*[contains(@id, 'ngdialog')]/div[2]/div[3]/div/button")
        close_btn.click()
    except:
        try:
            close_btn = driver.find_element(By.XPATH, "//*[contains(@id, 'ngdialog')]/div[2]/div[1]/div[3]/button")
            close_btn.click()
        except:
            driver.quit()
            driver.get("https://a810-dobnow.nyc.gov/publish/Index.html#!/")
            driver.delete_all_cookies()
            lic_btn.click()
            time.sleep(2)
            lic_type_btn.click()
            time.sleep(2)
            select.select_by_value('5')

#"C:\Users\ericl\PycharmProjects\CSCount\Copy of Active CS licensees 9-11-23.xlsx"
#Try to open Excel File. If Excel File doesn't exist or Filepath is wrong, raise error.
file_name = input("Please enter the Filepath of Excel File: ")
file_name.replace('\\','/')
print(file_name)
try:
    wb = openpyxl.load_workbook(file_name)
except:
    print("Invalid File Name: " + file_name)
    print("Exiting...")
    exit()

#Try to open Sheet. If name of Sheet doesn't exist, raise error.
sheet_name = input("Please enter the sheet name of the excel (Sheet1): ")
try:
    ws = wb[sheet_name]
except:
    print("Invalid Sheet Name: " + sheet_name)
    print("Exiting...")
    exit()

lic_col = int(input("Enter CS Licensee Column Number (Column A = 1): "))
starting_col = int(input("What column do you want to insert data, starting with Job Count? (22) "))
starting_row = int(input("What row do you want to start with? (2)"))

#Open a Chrome window with the url
driver = webdriver.Chrome()
driver.get("https://a810-dobnow.nyc.gov/publish/Index.html#!/")
driver.delete_all_cookies()

time.sleep(2)

# Click the Search By License button
lic_btn = driver.find_element(By.XPATH, value="//*[@id='content']/div[2]/div[1]/div[4]/div[2]/button")
lic_btn.click()
time.sleep(2)

# Click the Licensee Type drop down menu
lic_type_btn = driver.find_element(By.ID, value="LicSearchType")
lic_type_btn.click()
time.sleep(2)

# Select the type of license
select = Select(driver.find_element(By.ID, value="LicSearchType"))
select.select_by_value('5')

# Type Licensee Number into the search by license field
lic_search_box = driver.find_element(By.ID, value='LicLicenseNumbere')

for x in range(starting_row, ws.max_row):
    job_count = 0
    lic_num = ws.cell(starting_row, lic_col).value

    if ws.cell(starting_row, starting_col).value is None:

        # Fills License Search box with licensee number and searches
        lic_search_box.send_keys(lic_num)
        lic_search_box.send_keys(Keys.RETURN)
        time.sleep(2)

        # If there is no record, Total Job Count = 0
        # Example: License = 026836
        if "No records found for the given license number." in driver.page_source:
            print(str(starting_row) + " : No records")
            ws.cell(starting_row, starting_col).value = 0

        # Get the Total Job Count and Job Numbers and save it to Excel
        # Example: License = 026868
        elif "Associated Jobs with Active Permits" in driver.page_source:
            for i in range(1, 10):
                try:
                    job = driver.find_element(By.XPATH,
                                              "//*[contains(@id, 'ngdialog')]/div[2]/div[2]/div[2]/table/tbody/tr[" + str(
                                                  i) + "]/td[1]").text
                    starting_col += 1
                    ws.cell(starting_row, starting_col).value = job
                    i += 1
                    job_count += 1
                except:
                    print(str(starting_row) + " : " + str(job_count) + " records")
                    ws.cell(starting_row, column=22).value = job_count
                    break

        # Else Highlight Red

        wb.save(file_name)

        lic_search_box.clear()

        close_btn()

    starting_row += 1

print("Done")
time.sleep(5)
